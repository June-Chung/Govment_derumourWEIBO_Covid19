import time
import openpyxl as xl
from aip import AipNlp
import os


APP_ID = '19977603'
API_KEY = 'V29fofBiufPE1gLaNRjkNEGW'
SECRET_KEY = 'S7PeE3E9uawi1jVv49AZ6FGmxaUh3x7g'
client = AipNlp(APP_ID, API_KEY, SECRET_KEY)


contentset = set()
contentdic = {}


def dectiveEmotion(content):
    if len(content) > 1000:
        content = content[0:1000]
    if content in contentset:
        result = contentdic[content]
    else:
        contentset.add(content)
        print(content)
        time.sleep(0.5)
        result  = client.sentimentClassify(content)
        result = result['items'][0]['sentiment']
        contentdic[content] = result
    return result

def writeintoExcel(cc,flag):
    result_path = os.path.join('D:\Pythom-CODE\微热点情感识别', "data.xlsx")
    if os.path.exists(result_path):
        workbook = xl.load_workbook(result_path)
    else:
        workbook = xl.Workbook()
        workbook.save(result_path)
    sheet = workbook.active
    if flag == 0:
        sheet.append(['根微博ID','根微博用户ID','转发微博ID','转发微博用户ID','父微博ID','父微博用户ID','转发微博发布时间','转发微博内容','子微博','子微博@的人数','父微博','微博主题','情感'])
    else:
        for data in cc:
            sheet.append(data)
    workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** \n')


data = xl.load_workbook('情感识别数据/数据.xlsx')
sheet = data['Sheet1']

flag  =0
list4content = []
lasttheme = '江苏无锡高架坍塌，一侧翻遇难者为单亲爸爸'
currenttheme = ''
index = 1
for cell in sheet:
    if flag == 0:
        writeintoExcel([],0)
        flag =1
    else :
        currenttheme = cell[7].value
        if currenttheme!=lasttheme:
            contentdic.clear()
            contentset.clear()
        else:
            lasttheme = currenttheme
        print(index)
        index +=1
        list = [cell[0].value,cell[1].value,cell[2].value,cell[3].value,cell[4].value,cell[5].value,cell[6].value,cell[7].value,cell[8].value,cell[9].value,cell[10].value,cell[11].value]
        content = cell[8].value
        try:
            if content is None or content == '转发微博':
                content = cell[10].value
                if content is None or content == '转发微博':
                    content = cell[11].value
                    x = dectiveEmotion(content)
                    list.append(x)
                    list4content.append(list)
                    if len(list4content)>100:
                        writeintoExcel(list4content,1)
                        list4content = []
                else:
                    x = dectiveEmotion(content)
                    list.append(x)
                    list4content.append(list)
                    if len(list4content) > 100:
                        writeintoExcel(list4content, 1)
                        list4content = []
            else:
                x = dectiveEmotion(content)
                list.append(x)
                list4content.append(list)
                if len(list4content) > 100:
                    writeintoExcel(list4content, 1)
                    list4content = []
        except:
            if len(list4content)!=0:
                writeintoExcel(list4content,1)
                list4content = []
                time.sleep(15)




